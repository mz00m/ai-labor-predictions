#!/usr/bin/env python3
"""Join BLS AI exposure categories onto the occupation scorecard.

BLS blocks automated retrieval, so the XLSX must be downloaded by hand from
https://www.bls.gov/emp/publications/ai-exposure-categories.htm

Usage:
    python3 scripts/join-bls-ai-exposure.py <path-to-xlsx> [--apply]

Without --apply the script only reports what it would do.
"""
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
OCC_PATH = REPO / "src/data/enriched-occupations.json"

CATEGORIES = {"low", "moderate", "high", "very high"}
SOC_RE = re.compile(r"^\d{2}-\d{4}$")


def norm_soc(value):
    """O*NET 13-2011.00 and SOC 13-2011 both reduce to 13-2011."""
    if value is None:
        return None
    s = str(value).strip()
    m = re.match(r"^(\d{2}-\d{4})", s)
    return m.group(1) if m else None


def find_columns(rows):
    """Locate the header row plus the SOC and category column indexes.

    BLS table layouts shift between cycles, so detect rather than hardcode:
    the SOC column is the one whose cells look like SOC codes, and the
    category column is the one whose cells are the four category labels.
    """
    for header_idx, row in enumerate(rows[:30]):
        body = rows[header_idx + 1 : header_idx + 40]
        if not body:
            continue
        soc_col = cat_col = None
        for col in range(len(row)):
            vals = [r[col] for r in body if col < len(r) and r[col] is not None]
            if not vals:
                continue
            if soc_col is None and sum(1 for v in vals if norm_soc(v)) >= max(3, len(vals) * 0.6):
                soc_col = col
            if cat_col is None and sum(
                1 for v in vals if str(v).strip().lower() in CATEGORIES
            ) >= max(3, len(vals) * 0.6):
                cat_col = col
        if soc_col is not None and cat_col is not None:
            return header_idx, soc_col, cat_col
    return None, None, None


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    xlsx = Path(sys.argv[1])
    apply_changes = "--apply" in sys.argv
    if not xlsx.exists():
        sys.exit(f"XLSX not found: {xlsx}")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    best = (None, None, None, None)
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        header_idx, soc_col, cat_col = find_columns(rows)
        if soc_col is not None:
            best = (rows, header_idx, soc_col, cat_col)
            print(f"Using sheet '{ws.title}' (header row {header_idx + 1}, "
                  f"SOC col {soc_col}, category col {cat_col})")
            break
    rows, header_idx, soc_col, cat_col = best
    if rows is None:
        sys.exit("Could not locate SOC and AI-exposure-category columns. "
                 "Inspect the workbook and pass explicit columns.")

    bls = {}
    for row in rows[header_idx + 1:]:
        if soc_col >= len(row) or cat_col >= len(row):
            continue
        soc = norm_soc(row[soc_col])
        cat = row[cat_col]
        if soc and cat and str(cat).strip().lower() in CATEGORIES:
            bls[soc] = str(cat).strip()
    print(f"Parsed {len(bls)} occupations from BLS.")
    print("BLS category distribution:", dict(Counter(bls.values())))

    data = json.loads(OCC_PATH.read_text(encoding="utf-8"))
    occs = data["occupations"]
    matched = missed = 0
    for o in occs:
        soc = norm_soc(o.get("onetCode"))
        cat = bls.get(soc)
        if cat:
            o["blsAiExposure"] = cat
            o["blsAiExposureSource"] = "bls-ai-exposure-categories-2026"
            matched += 1
        else:
            missed += 1
    print(f"Matched {matched}/{len(occs)} scorecard occupations ({missed} unmatched).")

    # Cross-tab our 1-9 score against the BLS category to see whether they agree.
    order = ["Low", "Moderate", "High", "Very high"]
    buckets = {c: [] for c in order}
    for o in occs:
        c = o.get("blsAiExposure")
        if c in buckets and isinstance(o.get("exposure"), (int, float)):
            buckets[c].append(o["exposure"])
    print("\nOur 1-9 exposure score by BLS category:")
    for c in order:
        v = buckets[c]
        if v:
            print(f"  {c:<10} n={len(v):>3}  mean={sum(v)/len(v):.2f}  "
                  f"min={min(v)} max={max(v)}")

    if not apply_changes:
        print("\nDry run. Re-run with --apply to write enriched-occupations.json.")
        return
    OCC_PATH.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(f"\nWrote {OCC_PATH}")


if __name__ == "__main__":
    main()
